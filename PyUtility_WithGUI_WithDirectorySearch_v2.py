# -*- coding: utf-8 -*-
"""
Created on Fri Mar 20 16:37:16 2020

@author: Kshitij.Ranjan
"""

def refresh_excel(root_path,actuals_path):
    # Import packages
    import openpyxl
    from openpyxl.styles import Color, PatternFill, Font, Border
    from openpyxl.styles import colors
    from openpyxl.cell import Cell
    import datetime
    import dateutil.relativedelta
    import time
    import glob
    import os
    
    # Define the colour of the updated cells
    updatedColour = PatternFill(start_color='0000FF00',
                                end_color='0000FF00',
                                fill_type='solid')
    
    # Check run time
    start_time = time.time()
    # Read in the file to be used to update the original input file
    ActualsFile = glob.glob(actuals_path + "*.xlsx")
    ActualsFile = openpyxl.load_workbook(ActualsFile[0])
    actuals_sheet = ActualsFile.sheetnames
    Actuals_sheet1 = ActualsFile[actuals_sheet[0]]
    
    ### Reading in all the excels from a folder 
    # set all .xls files in your folder to list
    def getListOfFiles(dirName):
    # create a list of file and sub directories 
    # names in the given directory 
        listOfFile = os.listdir(dirName)
        allFiles = list()
        # Iterate over all the entries
        for entry in listOfFile:
            # Create full path
            fullPath = os.path.join(dirName, entry)
            # If entry is a directory then get the list of files in this directory 
            if os.path.isdir(fullPath):
                allFiles = allFiles + getListOfFiles(fullPath)
            else:
                if fullPath.endswith(".xlsx"):
                    allFiles.append(fullPath)
        return allFiles        
                
    allfiles = getListOfFiles(root_path)
    
    required_files = [];
    sum_1 = 0;
    for j in range(len(allfiles)):
        refFile = openpyxl.load_workbook(allfiles[j])
        allSheetNames = refFile.sheetnames
        for i in range(len(allSheetNames)):
            refsheet = refFile[allSheetNames[i]]
            try:
                if ((refsheet['B3'].value)+(refsheet['C3'].value)+(refsheet['D3'].value)+(refsheet['E3'].value)+
                    (refsheet['F3'].value)) in ('Marketing BucketsArticle Promotion Sub CategoryStore BrandingMetricCountry'):
                    sum_1 = sum_1 + 1

            except:
                pass
        if sum_1 >0:
            required_files.append(allfiles[j])
    for m in range(len(required_files)):
        theFile = openpyxl.load_workbook(required_files[m])
        allSheetNames = theFile.sheetnames
        excel_name = required_files[m].split('\\')
        excel_name = excel_name[len(excel_name)-1]
        path = required_files[m].replace(excel_name,"")
        path= path.replace("\\","/")
        os.chdir(path)
        current_workbook= m + 1
        total_wbs = len(required_files)
        print("Workbook ",  current_workbook , " of " , total_wbs)
        print("Workbook Name :" + excel_name)
        
        # Update values in all sheets for n-1 month
        for i in range(0,len(allSheetNames)):
            sheet1 = theFile[allSheetNames[i]]
            print("Sheet Name : " + allSheetNames[i])
            current_sheet = i+1
            total_sheets = len(allSheetNames) 
            print("Sheet " , current_sheet , "of"  ,total_sheets)
            print("")            
            # Iterate over the last 5 years
            for n in range(1,60):
                # Create the variable that stores the previous year month given current date
                x = datetime.datetime.now() - dateutil.relativedelta.relativedelta(months=n)
                col_to_update = str(x.year) + x.strftime("%m")

                # Identify the column (in the actuals file) that contains the data that needs to be updated
                def find_specific_col():
                    for column in "FGHIJKLMNOPQRSTUVWXYZ":
                        cell_name = "{}{}".format(column,1)
                        if Actuals_sheet1[cell_name].value == col_to_update:
                            col_name = column
                            return col_name
                # Store the column name that contains the data to be updated
                data_col = find_specific_col()

                # Get the column name that needs to be replaced in the Original input file
                col_to_update = str(x.year) + " " + x.strftime("%b")
                col_to_update

                # Identify the column to replace in the original input column
                def find_specific_cell():
                    for row in range(1, sheet1.max_row + 1):
                            for column in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":  # Here you can add or reduce the columns
                                cell_name = "{}{}".format(column, row)
                                if sheet1[cell_name].value == col_to_update:
                                    update_column = column
                                    return update_column
                update_column = find_specific_cell()

                # Run the rest of the code only if a column corresponding to this date exists 
                if data_col != None and update_column != None: 
                    # Get the Marketing Buckets,Article Promotion Sub Category,Store Branding and Metric from the original input file
                    def find_specific_cell():
                        for row in range(1, sheet1.max_row + 1):
                                for column in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":  # Here you can add or reduce the columns
                                    cell_name = "{}{}".format(column, row)
                                    if sheet1[cell_name].value == "Marketing Buckets":
                                        #print("{1} cell is located on {0}" .format(cell_name, currentSheet[cell_name].value))
                                        #print("cell position {} has value {}".format(cell_name, sheet1[cell_name].value))
                                        cell_name = "{}{}".format(column, row+1)
                                        #cell = sheet1.cell(row=row+1, column=column)
                                        return cell_name
                    Marketing_Bucket_cellname = find_specific_cell()
                    Marketing_Bucket = sheet1[Marketing_Bucket_cellname]
                    #print("Marketing Bucket is : " + Marketing_Bucket.value)

                    def find_specific_cell():
                        for row in range(1, sheet1.max_row + 1):
                                for column in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":  # Here you can add or reduce the columns
                                    cell_name = "{}{}".format(column, row)
                                    if sheet1[cell_name].value == "Article Promotion Sub Category":
                                        #print("{1} cell is located on {0}" .format(cell_name, currentSheet[cell_name].value))
                                        #print("cell position {} has value {}".format(cell_name, sheet1[cell_name].value))
                                        cell_name = "{}{}".format(column, row+1)
                                        #cell = sheet1.cell(row=row+1, column=column)
                                        return cell_name
                    ATK_Promotion_cellname = find_specific_cell()
                    ATK_Promotion = sheet1[ATK_Promotion_cellname]
                    #if ATK_Promotion.value != None :
                        #print("Article Promotion Sub Category is : " + ATK_Promotion.value)

                    def find_specific_cell():
                        for row in range(1, sheet1.max_row + 1):
                                for column in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":  # Here you can add or reduce the columns
                                    cell_name = "{}{}".format(column, row)
                                    if sheet1[cell_name].value == "Store Branding":
                                        #print("{1} cell is located on {0}" .format(cell_name, currentSheet[cell_name].value))
                                        #print("cell position {} has value {}".format(cell_name, sheet1[cell_name].value))
                                        cell_name = "{}{}".format(column, row+1)
                                        #cell = sheet1.cell(row=row+1, column=column)
                                        return cell_name
                    Store_Branding_cellname = find_specific_cell()
                    Store_Branding = sheet1[Store_Branding_cellname]
                    #print("Store Branding is : " + Store_Branding.value)

                    def find_specific_cell():
                        for row in range(1, sheet1.max_row + 1):
                                for column in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":  # Here you can add or reduce the columns
                                    cell_name = "{}{}".format(column, row)
                                    if sheet1[cell_name].value == "Metric":
                                        #print("{1} cell is located on {0}" .format(cell_name, currentSheet[cell_name].value))
                                        #print("cell position {} has value {}".format(cell_name, sheet1[cell_name].value))
                                        cell_name = "{}{}".format(column, row+1)
                                        #cell = sheet1.cell(row=row+1, column=column)
                                        return cell_name
                    Metric_cellname = find_specific_cell()
                    Metric = sheet1[Metric_cellname]
                    #print("Metric is : " + Metric.value)
                    #print(" ")

                    # Update values in the required column
                    def update_values():
                        for row in range(1, sheet1.max_row + 1):
                            for row1 in range(1, Actuals_sheet1.max_row + 1):
                                    MBucket_o = "{}{}".format('B', row)
                                    MBucket_a = "{}{}".format('A', row1)
                                    SubCat_o = "{}{}".format('C', row)
                                    SubCat_a = "{}{}".format('C', row1)
                                    Brand_o = "{}{}".format('D', row)
                                    Brand_a = "{}{}".format('B', row1)
                                    Metric_o = "{}{}".format('E', row)
                                    Metric_a = "{}{}".format('D', row1)
                                    country_o = "{}{}".format('F', row)
                                    country_a = "{}{}".format('E', row1)
                                    if sheet1[MBucket_o].value == Actuals_sheet1[MBucket_a].value:
                                        if Actuals_sheet1[SubCat_a].value != None:
                                            if sheet1[SubCat_o].value == Actuals_sheet1[SubCat_a].value:
                                                if sheet1[Brand_o].value == Actuals_sheet1[Brand_a].value:
                                                    if sheet1[Metric_o].value == Actuals_sheet1[Metric_a].value:
                                                        if sheet1[country_o].value == Actuals_sheet1[country_a].value:
                                                            cell_name_o = "{}{}".format(update_column, row)
                                                            cell_name_a = "{}{}".format(data_col, row1)
                                                            sheet1[cell_name_o].value = Actuals_sheet1[cell_name_a].value
                                                            sheet1[cell_name_o].fill = updatedColour
                                                        
                                        else:
                                            if sheet1[Brand_o].value == Actuals_sheet1[Brand_a].value:
                                                    if sheet1[Metric_o].value == Actuals_sheet1[Metric_a].value:
                                                        if sheet1[country_o].value == Actuals_sheet1[country_a].value:
                                                            cell_name_o = "{}{}".format(update_column, row)
                                                            cell_name_a = "{}{}".format(data_col, row1)
                                                            sheet1[cell_name_o].value = Actuals_sheet1[cell_name_a].value
                                                            sheet1[cell_name_o].fill = updatedColour
                                                        
                    # Run function
                    update_values()   
            theFile.save(excel_name)
    print("--- %s seconds ---" % (time.time() - start_time))


# root_path="C:\\Users\\Anupriya.John\\Documents\\Adidas\\Financial Forecasting\\Testing Folder\\"
# actuals_path="C:\\Users\\Anupriya.John\\Documents\\Adidas\\Financial Forecasting\\Testing Folder\\Actuals\\"
# refresh_excel(root_path,actuals_path)

import tkinter as tk

def update_entry_fields():
    
    root_path = e1.get()
    actuals_path = e2.get()
    print("Root Path: %s\nData Path: %s" % (root_path,actuals_path))
    refresh_excel(root_path,actuals_path)

master = tk.Tk()
tk.Label(master, 
         text="Enter the root path").grid(row=0)
tk.Label(master, 
         text="Enter the data path").grid(row=1)

e1 = tk.Entry(master)
e2 = tk.Entry(master)

e1.grid(row=0, column=1)
e2.grid(row=1, column=1)

tk.Button(master, 
          text='Click Here!', command=update_entry_fields).grid(row=3, 
                                                       column=1, 
                                                       sticky=tk.W, 
                                                       pady=4)

tk.mainloop()

